import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.conf import settings


WHISPER_MODEL = None
WHISPER_MODEL_SIZE = None
ONNX_MODEL = None


def get_whisper_model():
    global WHISPER_MODEL, WHISPER_MODEL_SIZE
    model_size = getattr(settings, 'WHISPER_MODEL_SIZE', 'base')
    device = getattr(settings, 'WHISPER_DEVICE', 'cpu')
    compute_type = getattr(settings, 'WHISPER_COMPUTE_TYPE', 'int8')

    if WHISPER_MODEL is None or WHISPER_MODEL_SIZE != model_size:
        WHISPER_MODEL = WhisperModel(model_size, device=device, compute_type=compute_type)
        WHISPER_MODEL_SIZE = model_size

    return WHISPER_MODEL


def get_onnx_model():
    global ONNX_MODEL
    if ONNX_MODEL is None:
        from onnx_asr import AutoModel
        ONNX_MODEL = AutoModel.from_pretrained("parakeet-tdt-0.6b-v3")
    return ONNX_MODEL


def transcribe_audio(audio_file, language='', model_size='base'):
    try:
        from faster_whisper import WhisperModel
        use_faster_whisper = True
    except ImportError:
        use_faster_whisper = False

    temp_path = None
    if hasattr(audio_file, 'temporary_file_path'):
        temp_path = audio_file.temporary_file_path()
    else:
        with tempfile.NamedTemporaryFile(delete=False, suffix='.wav') as tmp:
            for chunk in audio_file.chunks():
                tmp.write(chunk)
            temp_path = tmp.name

    if use_faster_whisper:
        model = get_whisper_model()
        segments, info = model.transcribe(temp_path, language=language or None, beam_size=5)

        transcript_parts = []
        segment_list = []
        for segment in segments:
            text = segment.text.strip()
            if text:
                transcript_parts.append(text)
                segment_list.append({
                    'start': segment.start,
                    'end': segment.end,
                    'text': text,
                })

        transcript = ' '.join(transcript_parts)
        duration = info.duration if info else 0.0
        detected_language = info.language if info else language or 'unknown'
        model_used = f'faster-whisper-{model_size}'
    else:
        model = get_onnx_model()
        result = model.transcribe(temp_path)
        transcript = result.text
        duration = result.duration if hasattr(result, 'duration') else 0.0
        detected_language = language or 'unknown'
        model_used = 'onnx-asr-parakeet'
        segment_list = []

    if temp_path and audio_file.name != temp_path:
        os.unlink(temp_path)

    return transcript, segment_list, duration, detected_language, model_used


def parse_soap_from_transcript(transcript):
    lines = transcript.strip().split('\n')
    text = ' '.join(lines)

    subjective = ''
    objective = ''
    assessment = ''
    plan = ''

    subject_markers = ['patient reports', 'patient complains', 'patient says', 'patient states', 'subjective']
    objective_markers = ['vitals', 'exam', 'physical exam', 'findings', 'objective', 'blood pressure', 'heart rate']
    assessment_markers = ['diagnosis', 'assessment', 'impression', 'likely', 'suspected']
    plan_markers = ['plan', 'treatment', 'prescribe', 'recommend', 'follow up', 'next steps']

    sentences = text.replace('?', '.').replace('!', '.').split('.')

    current_section = 'subjective'
    for sentence in sentences:
        s = sentence.strip().lower()
        if not s:
            continue

        if any(m in s for m in plan_markers):
            current_section = 'plan'
        elif any(m in s for m in assessment_markers):
            current_section = 'assessment'
        elif any(m in s for m in objective_markers):
            current_section = 'objective'
        elif any(m in s for m in subject_markers):
            current_section = 'subjective'

        if current_section == 'subjective':
            subjective += sentence.strip() + '. '
        elif current_section == 'objective':
            objective += sentence.strip() + '. '
        elif current_section == 'assessment':
            assessment += sentence.strip() + '. '
        elif current_section == 'plan':
            plan += sentence.strip() + '. '

    return {
        'subjective': subjective.strip(),
        'objective': objective.strip(),
        'assessment': assessment.strip(),
        'plan': plan.strip(),
    }


def generate_excel_report(report):
    wb = Workbook()

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    cell_alignment = Alignment(vertical='top', wrap_text=True)

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data_cell(cell):
        cell.border = thin_border
        cell.alignment = cell_alignment

    # Sheet 1: Patient Info
    ws_patient = wb.active
    ws_patient.title = 'Patient Info'
    patient_headers = ['Patient Name', 'Date of Birth', 'Gender', 'Medical Record Number', 'Phone', 'Email', 'Address']
    for col_idx, header in enumerate(patient_headers, 1):
        ws_patient.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_patient, 1, len(patient_headers))

    patient = report.patient
    patient_data = [
        patient.full_name,
        str(patient.date_of_birth),
        patient.get_gender_display(),
        patient.medical_record_number,
        patient.phone,
        patient.email,
        patient.address,
    ]
    for col_idx, value in enumerate(patient_data, 1):
        cell = ws_patient.cell(row=2, column=col_idx, value=value)
        style_data_cell(cell)

    for col_idx in range(1, len(patient_headers) + 1):
        ws_patient.column_dimensions[get_column_letter(col_idx)].width = 20

    # Sheet 2: Report
    ws_report = wb.create_sheet('Report')
    report_headers = ['Report Type', 'Consultation Date', 'Doctor', 'Subjective (S)', 'Objective (O)', 'Assessment (A)', 'Plan (P)', 'Raw Transcript']
    for col_idx, header in enumerate(report_headers, 1):
        ws_report.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_report, 1, len(report_headers))

    report_data = [
        report.get_report_type_display(),
        str(report.consultation_date),
        report.doctor.username,
        report.subjective,
        report.objective,
        report.assessment,
        report.plan,
        report.raw_transcript,
    ]
    for col_idx, value in enumerate(report_data, 1):
        cell = ws_report.cell(row=2, column=col_idx, value=value)
        style_data_cell(cell)

    col_widths = [18, 18, 15, 40, 40, 40, 40, 50]
    for col_idx, width in enumerate(col_widths, 1):
        ws_report.column_dimensions[get_column_letter(col_idx)].width = width

    # Sheet 3: Vitals (placeholder)
    ws_vitals = wb.create_sheet('Vitals')
    vitals_headers = ['Parameter', 'Value', 'Unit', 'Recorded At']
    for col_idx, header in enumerate(vitals_headers, 1):
        ws_vitals.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_vitals, 1, len(vitals_headers))

    for col_idx in range(1, len(vitals_headers) + 1):
        ws_vitals.column_dimensions[get_column_letter(col_idx)].width = 20

    # Sheet 4: Medications (placeholder)
    ws_meds = wb.create_sheet('Medications')
    meds_headers = ['Medication', 'Dosage', 'Frequency', 'Duration', 'Notes']
    for col_idx, header in enumerate(meds_headers, 1):
        ws_meds.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_meds, 1, len(meds_headers))

    for col_idx in range(1, len(meds_headers) + 1):
        ws_meds.column_dimensions[get_column_letter(col_idx)].width = 20

    # Save file
    media_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'excel')
    os.makedirs(media_dir, exist_ok=True)
    filename = f'report_{report.id}.xlsx'
    filepath = os.path.join(media_dir, filename)
    wb.save(filepath)

    return filepath


def generate_patient_excel(patient, reports):
    wb = Workbook()

    header_font = Font(bold=True, size=12, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    cell_alignment = Alignment(vertical='top', wrap_text=True)

    def style_header_row(ws, row_num, num_cols):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data_cell(cell):
        cell.border = thin_border
        cell.alignment = cell_alignment

    ws = wb.active
    ws.title = 'Patient Reports'
    headers = ['Report Type', 'Consultation Date', 'Doctor', 'Subjective', 'Objective', 'Assessment', 'Plan']
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)
    style_header_row(ws, 1, len(headers))

    for row_idx, report in enumerate(reports, 2):
        row_data = [
            report.get_report_type_display(),
            str(report.consultation_date),
            report.doctor.username,
            report.subjective[:200],
            report.objective[:200],
            report.assessment[:200],
            report.plan[:200],
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            style_data_cell(cell)

    for col_idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

    media_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'excel')
    os.makedirs(media_dir, exist_ok=True)
    filename = f'patient_{patient.id}_reports.xlsx'
    filepath = os.path.join(media_dir, filename)
    wb.save(filepath)

    return filepath